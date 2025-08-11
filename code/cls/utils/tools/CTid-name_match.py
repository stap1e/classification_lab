import pandas as pd
import os
from openpyxl import load_workbook

ctid_path = './new_data/radiomics_result_output.xlsx'
ctseg_path = "./new_data/ctseg-result_new.xlsx"
name_path1 = './new_data/0704数据-new1.xlsx'
name_path2 = './new_data/0704数据-new2.xlsx'
name_path3 = './new_data/0704数据-new3.xlsx'
old_t0_path = './datasets/time0_new.xlsx'
old_t1_path = './datasets/time1_new.xlsx'
old_ctseg_path = './datasets/CTSeg.xlsx'
need_ctid_path = './need_ctid_new.xlsx'

ctseg = pd.read_excel(ctseg_path)
ctid = pd.read_excel(ctid_path)
ctid_t0_old = pd.read_excel(old_t0_path, sheet_name=1)
ctid_t1_old = pd.read_excel(old_t1_path)
old_ctseg = pd.read_excel(old_ctseg_path)

# old data ctid
ctid_t0_old_name = ctid_t0_old['CTid'].tolist()
print(f"old data time0 num: {len(ctid_t0_old_name)}")
ctid_t1_old_name = ctid_t1_old['CTid'].tolist()
print(f"old data time1 num: {len(ctid_t1_old_name)}")
print(f"have name number: {len(ctid_t0_old_name) + len(ctid_t1_old_name) - len([x for x in ctid_t1_old_name if x in ctid_t0_old_name])}")

name1 = load_workbook(name_path1, data_only=True)
name1 = name1.worksheets[1]                       # sheet 1
visible_data1 = []
for row_idx, row in enumerate(name1.iter_rows(values_only=True), 1):
    if not name1.row_dimensions[row_idx].hidden:  # 检查行是否隐藏
        visible_data1.append(row)
name1 = pd.DataFrame(visible_data1[1:], columns=visible_data1[0]) if visible_data1 else pd.DataFrame()  

name2 = load_workbook(name_path2, data_only=True)
name2 = name2.worksheets[1]                       # sheet 1
visible_data2 = []
for row_idx, row in enumerate(name2.iter_rows(values_only=True), 1):
    if not name2.row_dimensions[row_idx].hidden:  # 检查行是否隐藏
        visible_data2.append(row)
name2 = pd.DataFrame(visible_data2[1:], columns=visible_data2[0]) if visible_data2 else pd.DataFrame()

name3 = load_workbook(name_path3, data_only=True)
name3 = name3.worksheets[1]                       # sheet 1
visible_data3 = []
for row_idx, row in enumerate(name3.iter_rows(values_only=True), 1):
    if not name3.row_dimensions[row_idx].hidden:  # 检查行是否隐藏
        visible_data3.append(row)
name3 = pd.DataFrame(visible_data3[1:], columns=visible_data3[0]) if visible_data3 else pd.DataFrame()
# prepare for name1,2,3 Nan, None
# name1 + name2 + name3 = 202 (excel)
name1 = name1[name1['姓名'].notna()]
name2 = name2[name2['姓名'].notna()]
name3 = name3[name3['姓名'].notna()]
print(f"origin excel have {len(name1) + len(name2) + len(name3)} patients...")

all_name = ctseg['name'].tolist()
all_ctid = ctseg['CTid'].tolist()
all_cpc = ctseg['CPC'].tolist()
print(f"all ctid and name length is: {len(all_ctid)}")     # 51
old_name = old_ctseg['name'].tolist()
old_ctid = old_ctseg['CTid'].tolist()
old_cttime  = old_ctseg['CTtime'].tolist()
print(f"old ctid and name length is: {len(old_ctid)}")     # 290

name1_n, name2_n, name3_n = name1['姓名'].tolist(), name2['姓名'].tolist(), name3['姓名'].tolist()
name1_cpc, name2_cpc, name3_cpc = name1['CPC分级'].tolist(), name2['CPC分级'].tolist(), name3['CPC分级'].tolist()
print(f"need to use name length is: 1-{len(name1_n)}, 2-{len(name2_n)}, 3-{len(name3_n)}")
ctid_n, ctid_cpc = ctid['CTid'].tolist(), ctid['CPC'].tolist()
print(f"ctid length is: {len(ctid_n)}")                    # 50

# find old ctid needed in time0,1,2
need_time1, need_time2, need_time3 = [], [], []
need_time1_id, need_time2_id, need_time3_id = [], [], []
nothave_1, nothave_2, nothave_3, nothave = [], [], [], []
nothave_ctid1, nothave_ctid2, nothave_ctid3, nothave_ctid = [], [], [], []
all_have_name, all_have_time, all_have_ctid = old_ctseg['name'].tolist(), old_ctseg['CTtime'].tolist(), old_ctseg['CTid'].tolist()
all_only_name = list(set(all_have_name))
print(f"old have name: {len(all_only_name)}")
for i in range(len(old_ctseg)):
    if all_have_name[i] in name1_n and all_have_time[i] == 0:
        need_time1.append(all_have_name[i])
        need_time1_id.append(all_have_ctid[i])
    if all_have_name[i] in name2_n and all_have_time[i] == 1:
        need_time2.append(all_have_name[i])
        need_time2_id.append(all_have_ctid[i])
    if all_have_name[i] in name3_n and all_have_time[i] == 2:
        need_time3.append(all_have_name[i])
        need_time3_id.append(all_have_ctid[i])

    if all_have_name[i] in name1_n and all_have_time[i] != 0 and all_have_name[i] not in nothave:
        nothave_1.append(all_have_name[i])
        nothave_ctid1.append(all_have_ctid[i])
        nothave.append(all_have_name[i])
        nothave_ctid.append(all_have_ctid[i])
    if all_have_name[i] in name2_n and all_have_time[i] != 1 and all_have_name[i] not in nothave:
        nothave_2.append(all_have_name[i])
        nothave_ctid2.append(all_have_ctid[i])
        nothave.append(all_have_name[i])
        nothave_ctid.append(all_have_ctid[i])
    if all_have_name[i] in name3_n and all_have_time[i] != 2 and all_have_name[i] not in nothave:
        nothave_3.append(all_have_name[i])
        nothave_ctid3.append(all_have_ctid[i])
        nothave.append(all_have_name[i])
        nothave_ctid.append(all_have_ctid[i])
print(f"old selected length is: 1-{len(need_time1)}, 2-{len(need_time2)}, 3-{len(need_time3)}")
print(f"old not have length is: 1-{len(nothave_1)}, 2-{len(nothave_2)}, 3-{len(nothave_3)}")
nothave_11 = [x for x in name1_n if x not in need_time1]
nothave_22 = [x for x in name2_n if x not in need_time2]
nothave_33 = [x for x in name3_n if x not in need_time3]
print(f"do not have name length: 1-{len(nothave_11)}, 2-{len(nothave_22)}, 3-{len(nothave_33)}")
# fine new ctid needed in time0,1,2
all_new_name, all_new_time, all_new_ctid = ctseg['name'], ctseg['CTtime'], ctseg['CTid']
all_only_name_new = list(set(all_new_name))
print(f"new have name: {len(all_only_name_new)}")
for i in range(len(ctseg)):
    if all_new_name[i] in name1_n and all_new_time[i] == 0:
        need_time1.append(all_new_name[i])
        need_time1_id.append(all_new_ctid[i])
    if all_new_name[i] in name2_n and all_new_time[i] == 1:
        need_time2.append(all_new_name[i])
        need_time2_id.append(all_new_ctid[i])
    if all_new_name[i] in name3_n and all_new_time[i] == 2:
        need_time3.append(all_new_name[i])
        need_time3_id.append(all_new_ctid[i])
print(f"new selected length is: 1-{len(need_time1)}, 2-{len(need_time2)}, 3-{len(need_time3)}")

# data1_path = './new_data/result_1_n.xlsx'
# data1_have = pd.read_excel(data1_path, sheet_name=1)
# data1_ctid_list = data1_have['CTid'].tolist()
# out_ctid = [x for x in need_time2_id if x not in data1_ctid_list]
nothave_1 = [x for x in name1_n if x not in need_time1]
nothave_2 = [x for x in name2_n if x not in need_time2]
nothave_3 = [x for x in name3_n if x not in need_time3]
# nothave_ctid11 = [x for x in name1_n if x not in need_time1]
# nothave_ctid22 = [x for x in name2_n if x not in need_time2]
# nothave_ctid33 = [x for x in name3_n if x not in need_time3]
print(f"do not have name length: 1-{len(nothave_11)}, 2-{len(nothave_22)}, 3-{len(nothave_33)}")

max_lt = max([len(need_time1), len(need_time2), len(need_time3)])
max_lt2 = max([len(need_time1_id), len(need_time2_id), len(need_time3_id)])
max_lt3 = max([len(nothave_1), len(nothave_2), len(nothave_3)])
max_l = max(max_lt, max_lt2, max_lt3)
print(f"old data needed num is: {len(need_time1) + len(need_time2) + len(need_time3)}")
print(f"old data_id needed num is: {len(need_time1_id) + len(need_time2_id) + len(need_time3_id)}")
print(f"old data needed num1 is: {len(need_time1)}, old data needed num2 is: {len(need_time2)}, old data needed num3 is: {len(need_time3)}")
for i in range(max_l - len(need_time1)):
    need_time1.append(None)
for i in range(max_l - len(need_time2)):
    need_time2.append(None)
for i in range(max_l - len(need_time3)):
    need_time3.append(None)
for i in range(max_l - len(need_time1_id)):
    need_time1_id.append(None)
for i in range(max_l - len(need_time2_id)):
    need_time2_id.append(None)
for i in range(max_l - len(need_time3_id)):
    need_time3_id.append(None)
for i in range(max_l - len(nothave_1)):
    nothave_1.append(None)
for i in range(max_l - len(nothave_2)):
    nothave_2.append(None)
for i in range(max_l - len(nothave_3)):
    nothave_3.append(None)
print(f"old data needed num1 is: {len(need_time1)}, old data needed num2 is: {len(need_time2)}, old data needed num3 is: {len(need_time3)}")
need_ctid_df = pd.DataFrame(need_time1, columns=["CTtime0_[0, 12)"])
need_ctid_df['CTtime1_[12, 72)'] = need_time2
need_ctid_df['CTtime2_[72, 336)'] = need_time3
need_ctid_df['CTtime0id_[0, 12)'] = need_time1_id
need_ctid_df['CTtime1id_[12, 72)'] = need_time2_id
need_ctid_df['CTtime2id_[72, 336)'] = need_time3_id
need_ctid_df['withoutCTid_[0, 12)'] = nothave_1
need_ctid_df['withoutCTid_[12, 72)'] = nothave_2
need_ctid_df['withoutCTid_[72, 336)'] = nothave_3
need_ctid_df.to_excel(need_ctid_path, index=False)


# true_name1, true_name_cpc1, true_ctid = [], [], []

